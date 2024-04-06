# contains helper functions
from openpyxl import load_workbook
from datetime import datetime

# parses nursing exam timetable
def nursing_exam_timetable_parser(file):

    # Define the list of column headers
    column_headers = ['Day', 'Campus', 'Coordinator', 'Courses', 'Hours', 
                      'Venue', 'Invigilators', 'Courses_Afternoon', 
                      'Hours_Afternoon', 'Invigilators_Afternoon', 
                      'Venue_Afternoon']

    def extract_course_info(column_data_dict, time_key, time_range):
        courses = []
        existing_course_codes = set()
        for i in range(len(column_data_dict["Day"])):
            course_code = column_data_dict[time_key][i]
            if course_code not in time_range and course_code not in existing_course_codes:
                course_info = {
                    "course_code": column_data_dict[time_key][i],
                    "coordinator": column_data_dict["Coordinator"][i],
                    "time": '8:30AM-11:30AM' if "8.30" in time_range[0] else '1:30PM-4:30PM',
                    "day": column_data_dict["Day"][i],
                    "campus": column_data_dict["Campus"][i],
                    "hrs": column_data_dict[f"Hours{'_Afternoon' if '_Afternoon' in time_key else ''}"][i],
                    "venue": column_data_dict[f"Venue{'_Afternoon' if '_Afternoon' in time_key else ''}"][i],
                    "invigilator": column_data_dict[f"Invigilators{'_Afternoon' if '_Afternoon' in time_key else ''}"][i]
                }
                
                courses.append(course_info)
                existing_course_codes.add(course_code)
        return courses

    # Loading the excel workbook
    wb_obj = load_workbook(file)
    sheet = wb_obj.active  # Activating the sheet for use

    # Initialize dictionary to store column data
    column_data_dict = {}

    # Iterate over columns and store data in dictionary
    for i, column in enumerate(sheet.iter_cols(values_only=True)):
        last_value = None
        column_data = []
        for cell in column:
            if cell is not None:
                # if column_headers[i] == 'Day':
                #     cell = cell.strftime('%A %d-%m-%Y')
                last_value = cell
                column_data.append(cell)
            else:
                column_data.append(last_value)
        if any(cell is not None for cell in column_data):
            column_data_dict[column_headers[i]] = column_data

    # Extract course information
    morning_exams = extract_course_info(column_data_dict, "Courses", ['8.30AM-11.30AM', '8.30-11.30 AM'])
    afternoon_exams = extract_course_info(column_data_dict, "Courses_Afternoon", ['1.30PM-4.30PM', '1.30-4.30PM'])

    # Combine morning and afternoon exams
    courses = morning_exams + afternoon_exams

    return courses


# parses nursing school timetable
def parse_nursing_timetable(file_path):
    # holds start and end times of every column number
    time_dictionary = {
        "2": ("8AM", "9AM"),
        "3": ("9AM", "10AM"),
        "4": ("10AM", "11AM"),
        "5": ("11AM", "12PM"),
        "6": ("12PM", "1PM"),
        "7": ("1PM", "2PM"),
        "8": ("2PM", "3PM"),
        "9": ("3PM", "4PM"),
        "10": ("4PM", "5PM"),
        "11": ("5PM", "6PM"),
    }

    days_of_the_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]
    unnecessary_course_values = ["CHAPEL", "CLP", "SDL", "KEY", 
                                "CLP-CLINICAL PRACTICE",
                                "SDL- SELF DIRECTED LEARNING"]

    courses = [] # will hold dictionaries of courses

    # loading the excel workbook
    wb_obj = load_workbook(filename=file_path)
    work_sheets = wb_obj.sheetnames # getting available work sheets

    # getting info from from first worksheet
    first_work_sheet = wb_obj[work_sheets[0]]

    # will hold course names and lecturers key being course code
    course_lectures = {}

    # iterating through the sheet
    for row in first_work_sheet.iter_rows(values_only=True):
        if row[1] == "CODE" or row[1] is None:
            continue

        # concantenating course name and the lecturer
        course_lectures[row[1]] = [row[2] , row[3]] 

     # getting info from from second worksheet
    second_work_sheet = wb_obj[work_sheets[1]]

    # course contents
    day = ""
    venue = ""
    course_time = ""
    course_name = ""

    # itearating the workbook
    for row_idx, rows in enumerate(second_work_sheet.iter_rows(values_only=True)):
        # skipping the unnecessary titles
        if row_idx in [0, 1, 2]:
            continue

        # skipping row with dates and taking the day of the course
        if rows[1] in days_of_the_week:
            day = rows[1]
            continue

        for idx ,row_value in enumerate(rows):
            # skipping the cohort part
            if idx == 1:
                continue

            # skipping unnecessary course values
            if row_value is None or row_value.strip() in unnecessary_course_values:
                continue

            # getting the venue
            if idx == 0:
                venue = row_value
                continue

            # getting course time
            start_time = time_dictionary[f"{idx}"][0]

            # setting course code
            course_name = row_value.strip()

            # handling merged rows
            remaining_cells = rows[idx+1:]
            last_none = 0
            for rem_idx, rem in enumerate(remaining_cells, start=idx):
                if rem is not None:
                    break

            # concatenating start time and end time
            course_time = start_time + "-" + time_dictionary[f"{rem_idx}"][1]
            
            courses.append({
                "course_code": course_name[:7],
                "lecturer": course_lectures[course_name[:7].strip()][1],
                "course_name": course_name,
                "day":day,
                "venue":venue,
                "time":course_time
                })
            
    return courses

# parses school exam timetable
def parse_school_exam_timetable(file):
    def time_difference(start_time, end_time): 
        """
        Returns the difference in hrs between two
        time intervals
        """

        format = '%I:%M%p'
        start_time = datetime.strptime(start_time, format)
        end_time = datetime.strptime(end_time, format)
        hrs = (end_time - start_time).total_seconds() / 3600
        return str(hrs)

    # loading the workbook
    wb_obj = load_workbook(file)

    work_sheets = wb_obj.sheetnames # getting available work sheets

    # getting info from from first workbook
    first_work_sheet = wb_obj[work_sheets[0]]

    # for sheet in book.worksheets: #For each worksheet


    rooms = {} # will hold room informationhold information about courses

    courses = [] # will hold courses information

    days_of_the_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

    # iterating through the rooms and storing them
    for column_one in first_work_sheet.iter_cols(values_only=True): #For each Column in a workshee
        for i, room in enumerate(column_one):
            # skipping none values and room word
            if room is None or room == "ROOM":
                continue
            rooms[f'{i}'] = room
        break
    # a list of the remaining columns without the rooms
    data_columns = list(first_work_sheet.iter_cols(values_only=True))[1:]

    # values for the course code
    day = ""
    time = ""
    course_code = ""

    # iterating through the columns data
    for column in data_columns:
        # SKIPPING TUESDAY CHAPEL TIME
        if column[1] is not None and column[1].split(" ")[0] == "TUESDAY":
            if column[2].strip() == "8:30AM-9:30AM":
                # set the day before skipping the columns
                day = column[1]
                continue
        
        # SKIPPING THURSDAY CHAPEL TIME
        if column[1] is not None and column[1].split(" ")[0] == "THURSDAY":
            if column[2].strip() == "8:30AM-9:30AM":
                # set the day before skipping the columns
                day = column[1]
                continue
        
        for idx, value in enumerate(column):
            # skipping empty cell values
            if value is None:
                continue

            # checking if its date and day specification
            if value.split(" ")[0] in days_of_the_week:
                day = value

            # checking if its time specification
            elif value[0].isdigit():
                course_time = value.strip()
                start_time = course_time.split("-")[0]
                end_time = course_time.split("-")[1]
                hours = time_difference(start_time, end_time)
            else:
                course_code = value
            
                courses.append(
                    {
                        "course_code": course_code,
                        "day": day,
                        "time": course_time,
                        "venue": rooms[f'{idx}'],
                        # handles errors recorded in exam timetable defaults to 2
                        "hrs": "2" if hours[0] == "-" else hours[0],
                    }
                )
    return courses