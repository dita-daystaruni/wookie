from openpyxl import load_workbook
from datetime import datetime

course_info = {
    "Course_Name": "",
    "Coordinator": "",
    "Day": "",
    "Time": "",
    "Campus": "",
    "Hrs": "",
    "Venue": "",
    "Invigilator": ""
    }

courses = [] # will hold dictionaries of courses
# loading the excel workbook
wb_obj = load_workbook(filename="Nursing_TimeTable.xlsx")
sheet = wb_obj.active # activating the sheet for use

# itearating the workbook
for row in sheet.iter_rows(values_only=True):
    # determining the header columns
    # if type(row[0]) is datetime:
    #     print("header") 
    for value in row:
        print(value)
