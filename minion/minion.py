"""
Minion Bot.

Author: Erick Muuo
Description:

The minion bot parses excel files, for timetable and other stuff

Copyright: (C) Academia 2024
"""

import openpyxl


class Minion:
    """Minion.

    Defines funitionality to parse excel files.
    """

    def __init__(self, file: str):
        """Construct a minion instance that parses the excel file."""
        self.file = file

    def is_blankRow(self, columns: iter) -> bool:
        for cell in columns:
            # print(cell)
            if cell.value is not None:
                return False
        return True

    def digest(self):
        """Digest the excel file and return a json object."""
        try:
            # loading the workbook
            wb_obj = openpyxl.load_workbook(self.file)

            work_sheets = wb_obj.sheetnames # getting available work sheets

            # getting info from from first workbook
            first_work_sheet = wb_obj[work_sheets[0]]

            # for sheet in book.worksheets: #For each worksheet


            rooms = {} # will hold room informationhold information about courses

            courses = [] # will hold courses information

            days_of_the_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

            # iterating through the rooms and storing them
            for column_one in first_work_sheet.iter_cols(values_only=True): #For each Column in a workshee
                for i, room in enumerate(column_one):
                    # skipping none values and room word
                    if room is None or room == "ROOM":
                        continue
                    rooms[f'{i}'] = room
                break
            # a list of the remaining columns without the rooms
            data_columns = list(first_work_sheet.iter_cols(values_only=True))[1:]

            # values for the course code
            day = ""
            time = ""
            course_code = ""

            # iterating through the columns data
            for column in data_columns:
                # SKIPPING TUESDAY CHAPEL TIME
                if column[1] is not None and column[1].split(" ")[0] == "TUESDAY":
                    if column[2].strip() == "8:30AM-9:30AM":
                        continue
                
                # SKIPPING THURSDAY CHAPEL TIME
                if column[1] is not None and column[1].split(" ")[0] == "THURSDAY":
                    if column[2].strip() == "8:30AM-9:30AM":
                        continue
                
                for idx, value in enumerate(column):
                    # skipping empty cell values
                    if value is None:
                        continue

                    # checking if its date and day specification
                    if value.split(" ")[0] in days_of_the_week:
                        day = value

                    # checking if its time specification
                    elif value[0].isdigit():
                        time = value
                    else:
                        course_code = value
                    
                        courses.append(
                            {
                                "course_code": course_code,
                                "day": day,
                                "time": time,
                                "room": rooms[f'{idx}']
                            }
                        )
            return courses
        except Exception as e:
            print(e)
